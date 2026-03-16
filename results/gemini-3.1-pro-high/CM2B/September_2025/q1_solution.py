import os
import math
import statistics
from openpyxl import load_workbook, Workbook

SOURCE = "exams/CM2B/September_2025/Answer-Booklet.xlsx"
wb = load_workbook(SOURCE, data_only=True)
ws = wb["Q1 Data"]

returns_A = []
returns_B = []

for row in ws.iter_rows(min_row=6, max_row=1005, values_only=True):
    if row[1] is not None:
        returns_A.append(row[2])
        returns_B.append(row[3])

value_A = [100 * (1 + r) for r in returns_A]
value_B = [100 * (1 + r) for r in returns_B]

print("=== Q1 Results ===")

# (i) Calculate the mean and variance of the value of each asset at time t = 1.
mean_V_A = statistics.mean(value_A)
var_V_A = statistics.variance(value_A)
mean_V_B = statistics.mean(value_B)
var_V_B = statistics.variance(value_B)

print(f"(i) Asset A: Mean Value = {mean_V_A:.4f}, Variance = {var_V_A:.4f}")
print(f"(i) Asset B: Mean Value = {mean_V_B:.4f}, Variance = {var_V_B:.4f}")

# (ii) Calculate the expected utility to the investor of each asset at time t = 1.
# U(w) = - 1 / (500 * w^2)
def utility(w):
    return -1 / (500 * (w ** 2))

util_A = [utility(v) for v in value_A]
util_B = [utility(v) for v in value_B]

exp_util_A = statistics.mean(util_A)
exp_util_B = statistics.mean(util_B)

print(f"(ii) Asset A: Expected Utility = {exp_util_A:.8f}")
print(f"(ii) Asset B: Expected Utility = {exp_util_B:.8f}")

# (iv) Calculate the 99% Value at Risk (VaR) of the value of each asset at time t = 1.
# Sort values
value_A_sorted = sorted(value_A)
value_B_sorted = sorted(value_B)

# Index for 1st percentile: 1000 * 0.01 = 10. So 10th lowest value (index 9, or interpolate)
# Excel PERCENTILE.EXC usually interpolates. Let's compute simply with index
# 99% VaR is the 1st percentile of the value distribution
# Usually 99% VaR = V_0 - 1st percentile or just the loss
# Since 1% out of 1000 is exactly 10, let's use the 10th value (index 9)
p1_A = value_A_sorted[int(1000 * 0.01) - 1]
p1_B = value_B_sorted[int(1000 * 0.01) - 1]

# What does numpy percentile say for 1% of 1000 items?
import math
def int_pctl(data, p):
    # simplest percentile estimating method
    data.sort()
    n = len(data)
    i = p * (n - 1)
    f = math.floor(i)
    c = math.ceil(i)
    if f == c:
        return data[int(i)]
    d0 = data[int(f)] * (c - i)
    d1 = data[int(c)] * (i - f)
    return d0 + d1

v_1_A = int_pctl(value_A, 0.01)
v_1_B = int_pctl(value_B, 0.01)

print(f"(iv) Asset A: 1st percentile value = {v_1_A:.4f}, VaR from 100 = {100-v_1_A:.4f}")
print(f"(iv) Asset B: 1st percentile value = {v_1_B:.4f}, VaR from 100 = {100-v_1_B:.4f}")

# (v) Calculate the expected shortfall of the value of each asset at time t = 1 relative to $150.
# ES = E[max(150 - V, 0)]
es_A = sum([max(150 - v, 0) for v in value_A]) / len(value_A)
es_B = sum([max(150 - v, 0) for v in value_B]) / len(value_B)

print(f"(v) Asset A: Expected Shortfall rel to $150 = {es_A:.4f}")
print(f"(v) Asset B: Expected Shortfall rel to $150 = {es_B:.4f}")

# (vii) Derive, using your answer to part (i), estimates of alpha and beta for each asset.
# Mean of Beta: mu_R = alpha / (alpha + beta)
# Var of Beta: var_R = (alpha * beta) / ((alpha + beta)^2 * (alpha + beta + 1))
# Let m = mean_R, v = var_R. Then alpha = m * (m*(1-m)/v - 1)
mean_R_A = statistics.mean(returns_A)
var_R_A = statistics.variance(returns_A)
mean_R_B = statistics.mean(returns_B)
var_R_B = statistics.variance(returns_B)

print(f"Returns A: Mean = {mean_R_A:.4f}, Var = {var_R_A:.4f}")
print(f"Returns B: Mean = {mean_R_B:.4f}, Var = {var_R_B:.4f}")

alpha_A = mean_R_A * (mean_R_A * (1 - mean_R_A) / var_R_A - 1)
beta_A = (1 - mean_R_A) * (mean_R_A * (1 - mean_R_A) / var_R_A - 1)

alpha_B = mean_R_B * (mean_R_B * (1 - mean_R_B) / var_R_B - 1)
beta_B = (1 - mean_R_B) * (mean_R_B * (1 - mean_R_B) / var_R_B - 1)

print(f"(vii) Asset A: alpha = {alpha_A:.4f}, beta = {beta_A:.4f}")
print(f"(vii) Asset B: alpha = {alpha_B:.4f}, beta = {beta_B:.4f}")

# Save to workings.xlsx
out_wb = Workbook()
out_ws = out_wb.active
out_ws.title = "Q1 Workings"

out_ws.append(["Part", "Metric", "Asset A", "Asset B"])
out_ws.append(["(i)", "Mean Value", mean_V_A, mean_V_B])
out_ws.append(["(i)", "Variance Value", var_V_A, var_V_B])
out_ws.append(["(ii)", "Expected Utility", exp_util_A, exp_util_B])
out_ws.append(["(iv)", "1st Percentile Value", v_1_A, v_1_B])
out_ws.append(["(v)", "Expected Shortfall rel to $150", es_A, es_B])
out_ws.append(["(vii)", "Return Mean", mean_R_A, mean_R_B])
out_ws.append(["(vii)", "Return Var", var_R_A, var_R_B])
out_ws.append(["(vii)", "Alpha", alpha_A, alpha_B])
out_ws.append(["(vii)", "Beta", beta_A, beta_B])

if not os.path.exists("results/gemini-3.1-pro-high/CM2B/September_2025/workings.xlsx"):
    out_wb.save("results/gemini-3.1-pro-high/CM2B/September_2025/workings.xlsx")

import json
results = {
    "A": {
        "mean_V": mean_V_A, "var_V": var_V_A, "exp_u": exp_util_A, "P1_V": v_1_A, "VaR": 100-v_1_A, "ES": es_A, "alpha": alpha_A, "beta": beta_A
    },
    "B": {
        "mean_V": mean_V_B, "var_V": var_V_B, "exp_u": exp_util_B, "P1_V": v_1_B, "VaR": 100-v_1_B, "ES": es_B, "alpha": alpha_B, "beta": beta_B
    }
}
print(json.dumps(results, indent=2))
