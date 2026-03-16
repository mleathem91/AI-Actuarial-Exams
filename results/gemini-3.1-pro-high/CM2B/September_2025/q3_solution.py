import math

S = 40000
K = 39000
T = 5
r = 0.025
P_market = 2000

# Normal fully python helper
def norm_cdf(x):
    return (1.0 + math.erf(x / math.sqrt(2.0))) / 2.0

def norm_pdf(x):
    return math.exp(-0.5 * x**2) / math.sqrt(2 * math.pi)

def d1(s_val, sigma_val):
    if sigma_val <= 0: return 0
    return (math.log(s_val / K) + (r + 0.5 * sigma_val**2) * T) / (sigma_val * math.sqrt(T))

def d2(s_val, sigma_val):
    if sigma_val <= 0: return 0
    return d1(s_val, sigma_val) - sigma_val * math.sqrt(T)

def bs_put(s_val, sigma_val):
    d1_val = d1(s_val, sigma_val)
    d2_val = d2(s_val, sigma_val)
    return K * math.exp(-r * T) * norm_cdf(-d2_val) - s_val * norm_cdf(-d1_val)

def bs_vega(s_val, sigma_val):
    d1_val = d1(s_val, sigma_val)
    return s_val * math.sqrt(T) * norm_pdf(d1_val)

# (iii) Implied Volatility
# Using Newton Raphson
sigma_est = 0.2
for _ in range(50):
    put = bs_put(S, sigma_est)
    vega = bs_vega(S, sigma_est)
    diff = put - P_market
    if abs(diff) < 1e-4:
        break
    if vega == 0:
        break
    sigma_est -= diff / vega

iv = sigma_est
print(f"(iii) Implied Volatility: {iv:.6f} (or {iv*100:.2f}%)")

# (iv) Estimate value of Vega for the put option
# Vega is the same for Call and Put
vega_est = bs_vega(S, iv)
print(f"(iv) Vega: {vega_est:.4f}")

# (v) Calculate updated value of put option for volatilities between 5% and 15%
print("(v) Option Values for sigmas between 5% and 15%:")
sigmas = [0.05, 0.06, 0.07, 0.08, 0.09, 0.10, 0.11, 0.12, 0.13, 0.14, 0.15]
results_v = []
for s in sigmas:
    # (a) Approx using Vega
    # P(s) = P(iv) + Vega * (s - iv)
    # Wait, the question implies using Vega we estimate it, P_market is 2000, but Vega is per 100% or per 1%?
    # Standard Vega is vega_est, delta sigma = s - iv
    approx = P_market + vega_est * (s - iv)
    
    # (b) Exact using BS
    exact = bs_put(S, s)
    results_v.append((s, approx, exact))
    print(f"Sigma = {s*100:.0f}% : Approx = {approx:.4f}, Exact = {exact:.4f}")

# Generate Chart via QuickChart
import urllib.request
import urllib.parse
import json

chart_config = {
    "type": "line",
    "data": {
        "labels": [f"{s*100:.0f}%" for s in sigmas],
        "datasets": [
            {
                "label": "Approx by Vega",
                "data": [r[1] for r in results_v],
                "borderColor": "blue",
                "fill": False
            },
            {
                "label": "Exact BS",
                "data": [r[2] for r in results_v],
                "borderColor": "red",
                "fill": False
            }
        ]
    },
    "options": {
        "title": {
            "display": True,
            "text": "Put Option Prices: Approx vs BS Model"
        }
    }
}
encoded_config = urllib.parse.quote(json.dumps(chart_config))
url = f"https://quickchart.io/chart?c={encoded_config}"
try:
    urllib.request.urlretrieve(url, "results/gemini-3.1-pro-high/CM2B/September_2025/q3_plot.png")
    print("Plot saved to q3_plot.png")
except Exception as e:
    print("Failed to save plot", e)

# Save to workings.xlsx
from openpyxl import load_workbook
wb = load_workbook("results/gemini-3.1-pro-high/CM2B/September_2025/workings.xlsx")
ws = wb.create_sheet("Q3 Workings")
ws.append(["Q3 Results"])
ws.append(["(i) r", r])
c_calc = P_market + S - K * math.exp(-r * T)
ws.append(["(ii) Call Value", c_calc])
ws.append(["(iii) Implied Vol", iv])
ws.append(["(iv) Vega", vega_est])
ws.append([])
ws.append(["Sigma", "Approx Put", "Exact Put"])
for s, app, ex in results_v:
    ws.append([s, app, ex])
wb.save("results/gemini-3.1-pro-high/CM2B/September_2025/workings.xlsx")
