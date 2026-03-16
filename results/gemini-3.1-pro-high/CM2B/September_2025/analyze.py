import os
import openpyxl

os.makedirs('results/gemini-3.1-pro-high/CM2B/September_2025', exist_ok=True)

wb = openpyxl.load_workbook('exams/CM2B/September_2025/Answer-Booklet.xlsx', data_only=True)

with open('results/gemini-3.1-pro-high/CM2B/September_2025/analyze.txt', 'w') as f:
    for sheet in ['Q1 Data', 'Q2 Data', 'Q3 Data']:
        f.write(f"--- {sheet} ---\n")
        ws = wb[sheet] if sheet in wb.sheetnames else None
        if ws is None:
            f.write("Not found\n")
            continue
        for row in ws.iter_rows(min_row=10, max_row=25, values_only=True):
            f.write(str(row) + '\n')
