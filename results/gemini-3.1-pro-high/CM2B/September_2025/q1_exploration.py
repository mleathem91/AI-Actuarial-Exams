import os
import math
import statistics
from openpyxl import load_workbook, Workbook

SOURCE = "exams/CM2B/September_2025/Answer-Booklet.xlsx"
wb = load_workbook(SOURCE, data_only=True)
ws = wb["Q1 Data"]

returns_A = []
returns_B = []

for row in ws.iter_rows(min_row=6, max_row=1005, values_only=True):
    if row[1] is not None:
        returns_A.append(row[2])
        returns_B.append(row[3])

# Value at t=1: V = 100 * (1 + R)
# Wait, let's print max and min to see if they are between 0 and 1
print("Max return A:", max(returns_A), "Min return A:", min(returns_A))
print("Max return B:", max(returns_B), "Min return B:", min(returns_B))
