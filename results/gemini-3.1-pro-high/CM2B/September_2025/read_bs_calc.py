import os
from openpyxl import load_workbook

SOURCE = "exams/CM2B/September_2025/Answer-Booklet.xlsx"
wb = load_workbook(SOURCE, data_only=True)
ws = wb["B-S calculator"]

print("--- B-S calculator ---")
for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
    print(row)
