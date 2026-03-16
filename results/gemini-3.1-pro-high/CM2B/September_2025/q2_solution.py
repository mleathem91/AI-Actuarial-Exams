import math
import urllib.request
import urllib.parse
import json
from openpyxl import load_workbook

kappa = 0.7
theta = 0.03
sigma = 0.09
r0 = 0.03

def normal_cdf_inv(p):
    # Approximation of inverse normal CDF
    # Using Beasley-Springer-Moro or similar simple approximation
    a1 = -39.69683028665376
    a2 = 220.9460984245205
    a3 = -275.9285104469687
    a4 = 138.3577518672690
    a5 = -30.66479806614716
    a6 = 2.506628277459239
    
    b1 = -54.47609879822406
    b2 = 161.5858368580409
    b3 = -155.6989798598866
    b4 = 66.80131188771972
    b5 = -13.28068155288572
    
    c1 = -0.007784894002430293
    c2 = -0.3223964580411365
    c3 = -2.400758277161838
    c4 = -2.549732539343734
    c5 = 4.374664141464968
    c6 = 2.938163982698783
    
    d1 = 0.007784695709041462
    d2 = 0.3224671290700398
    d3 = 2.445134137142996
    d4 = 3.754408661907416
    
    q = p - 0.5
    if abs(q) <= 0.425:
        r = 0.180625 - q * q
        val = q * (((((a1 * r + a2) * r + a3) * r + a4) * r + a5) * r + a6) / \
              (((((b1 * r + b2) * r + b3) * r + b4) * r + b5) * r + 1)
        return val
    else:
        r = p if q < 0 else 1 - p
        if r <= 0: return -1e9
        r = math.sqrt(-math.log(r))
        val = (((((c1 * r + c2) * r + c3) * r + c4) * r + c5) * r + c6) / \
              ((((d1 * r + d2) * r + d3) * r + d4) * r + 1)
        return val if q < 0 else -val

U = [
    [0.91, 0.38, 0.461, 0.711, 0.327],
    [0.574, 0.229, 0.718, 0.083, 0.685],
    [0.075, 0.636, 0.289, 0.396, 0.451]
]

paths = []
z_vals = []
for i in range(3):
    r_path = [r0]
    z_path = []
    for j in range(5):
        u = U[i][j]
        z = normal_cdf_inv(u)
        z_path.append(z)
        r_t = r_path[-1]
        dr = kappa * (theta - r_t) + sigma * math.sqrt(r_t) * z
        r_next = max(0, r_t + dr) 
        r_path.append(r_next)
    paths.append(r_path)
    z_vals.append(z_path)

print("=== Q2 Results ===")
for i, path in enumerate(paths):
    print(f"Path {i+1} : {[round(x, 4) for x in path]}")
    print(f"   Z : {[round(z, 4) for z in z_vals[i]]}")

# Generate Chart via QuickChart
chart_config = {
    "type": "line",
    "data": {
        "labels": ["0", "1", "2", "3", "4", "5"],
        "datasets": [
            {
                "label": f"Path 1",
                "data": paths[0],
                "borderColor": "red",
                "fill": False
            },
            {
                "label": f"Path 2",
                "data": paths[1],
                "borderColor": "green",
                "fill": False
            },
            {
                "label": f"Path 3",
                "data": paths[2],
                "borderColor": "blue",
                "fill": False
            }
        ]
    },
    "options": {
        "title": {
            "display": True,
            "text": "Simulated 1-year Short Rate (r_t) over 5 Years (CIR Model)"
        },
        "scales": {
            "yAxes": [{
                "scaleLabel": {
                    "display": True,
                    "labelString": "Short Rate"
                }
            }],
            "xAxes": [{
                "scaleLabel": {
                    "display": True,
                    "labelString": "Time (Years)"
                }
            }]
        }
    }
}
encoded_config = urllib.parse.quote(json.dumps(chart_config))
url = f"https://quickchart.io/chart?c={encoded_config}"
urllib.request.urlretrieve(url, "results/gemini-3.1-pro-high/CM2B/September_2025/q2_plot.png")
print("Plot saved to q2_plot.png")

# (i) Calculate
def B(t, T, k):
    return (1 - math.exp(-k * (T - t))) / k

def A(t, T, k, th, s):
    B_val = B(t, T, k)
    term1 = (th - (s**2)/(2 * k**2)) * (B_val - (T - t))
    term2 = (s**2 * B_val**2) / (4 * k)
    return math.exp(term1 - term2)

def P(t, T, r, k, th, s):
    return A(t, T, k, th, s) * math.exp(-B(t, T, k) * r)

P_0_1 = P(0, 1, r0, kappa, theta, sigma)
price_X = 100 * P_0_1
spot_1 = -math.log(P_0_1) / 1
print(f"(i)(a) Price of Bond X: {price_X:.4f}")
print(f"(i)(b) 1-year spot rate: {spot_1:.4f} (or {spot_1*100:.2f}%)")

# (ii) Calculate
price_Y = 92
spot_3 = -math.log(price_Y / 100) / 3
P_0_3 = price_Y / 100
P_0_1_bond = P_0_1
f_1_3 = -math.log(P_0_3 / P_0_1_bond) / 2
print(f"(ii)(a) 3-year spot rate: {spot_3:.4f} (or {spot_3*100:.2f}%)")
print(f"(ii)(b) 2-year forward rate at t=1: {f_1_3:.4f} (or {f_1_3*100:.2f}%)")

# Saving to Excel workings
try:
    wb = load_workbook("results/gemini-3.1-pro-high/CM2B/September_2025/workings.xlsx")
    if "Q2 Data" not in wb.sheetnames:
        ws = wb.create_sheet("Q2 Workings")
    else:
        ws = wb["Q2 Workings"]
        
    ws.append(["Q2 Results"])
    ws.append(["Path", "t=0", "t=1", "t=2", "t=3", "t=4", "t=5"])
    for i, p in enumerate(paths):
        ws.append([f"Path {i+1} r"] + p)
        ws.append([f"Path {i+1} Z", "N/A"] + z_vals[i])
        ws.append([f"Path {i+1} U", "N/A"] + U[i])
    wb.save("results/gemini-3.1-pro-high/CM2B/September_2025/workings.xlsx")
except Exception as e:
    print("Could not update workings", e)
