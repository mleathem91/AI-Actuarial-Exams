import openpyxl

wb = openpyxl.load_workbook("exams/CM1B/September_2025/Answer-Booklet.xlsx", data_only=True)
ws = wb["Q2 Base"]

m_term_q = {}
f_term_q = {}
m_ann_q = {}
f_ann_q = {}

for r in range(6, 120):
    v_age = ws.cell(row=r, column=1).value
    if v_age is not None and isinstance(v_age, (int, float)):
        age = int(v_age)
        m_term_q[age] = float(ws.cell(row=r, column=2).value)
        f_term_q[age] = float(ws.cell(row=r, column=3).value)
        
    v_age_ann = ws.cell(row=r, column=5).value
    if v_age_ann is not None and isinstance(v_age_ann, (int, float)):
        age = int(v_age_ann)
        m_ann_q[age] = float(ws.cell(row=r, column=6).value)
        f_ann_q[age] = float(ws.cell(row=r, column=7).value)

age_m = 30
age_f = 25
term = 35
sum_assured = 100000
annuity_base = 12000
rev_prop = 0.5
i_def = 0.04
i_pay = 0.05
v_def = 1 / (1 + i_def)
v_pay = 1 / (1 + i_pay)

ev_term = 0.0
p_joint_t_minus_1 = 1.0
results_term = []

for t_yr in range(1, term + 1):
    q_m = m_term_q.get(age_m + t_yr - 1, 1.0)
    q_f = f_term_q.get(age_f + t_yr - 1, 1.0)
    p_m_surv = 1 - q_m
    p_f_surv = 1 - q_f
    p_joint_surv = p_m_surv * p_f_surv
    prob_fails = 1 - p_joint_surv
    term_cf = sum_assured * (v_def ** t_yr) * p_joint_t_minus_1 * prob_fails
    ev_term += term_cf
    results_term.append({
        "t": t_yr, "age_m": age_m + t_yr - 1, "age_f": age_f + t_yr - 1,
        "p_joint_start": p_joint_t_minus_1, "prob_fails": prob_fails, "cf_pv": term_cf
    })
    p_joint_t_minus_1 *= p_joint_surv

prob_both_survive_to_vesting = p_joint_t_minus_1
ev_annuity_at_vesting = 0.0
for t in range(5): ev_annuity_at_vesting += annuity_base * (v_pay ** t)

ev_ann_post_guar = 0.0
p_m_t = [1.0]
p_f_t = [1.0]
max_years = max(max(m_ann_q.keys()) - 65, max(f_ann_q.keys()) - 60) + 2

for t in range(max_years):
    if t > 0:
        p_m_t.append(p_m_t[-1] * (1 - m_ann_q.get(65 + t - 1, 1.0)))
        p_f_t.append(p_f_t[-1] * (1 - f_ann_q.get(60 + t - 1, 1.0)))

for t in range(5, max_years):
    pv_factor = v_pay ** t
    expected_payment = 6000 * p_m_t[t] + 6000 * p_f_t[t]
    ev_ann_post_guar += expected_payment * pv_factor

annuity_value_at_65_M = sum(annuity_base * (v_pay ** t) for t in range(5)) + ev_ann_post_guar
ev_annuity_total = prob_both_survive_to_vesting * (v_def ** term) * annuity_value_at_65_M

with open("results/gemini-3.1-pro-high/CM1B/September_2025/q2_output.txt", "w", encoding="utf-8") as f:
    f.write(f"(i) Expected value of joint life term assurance: {ev_term:.4f}\n")
    f.write(f"Prob both survive to vesting = {prob_both_survive_to_vesting:.6f}\n")
    f.write(f"Value of annuity at vesting M65/F60: {annuity_value_at_65_M:.4f}\n")
    f.write(f"(ii) Expected value of deferred annuity at issue: {ev_annuity_total:.4f}\n")

from openpyxl import Workbook
import os
OUT_DIR = "results/gemini-3.1-pro-high/CM1B/September_2025"
os.makedirs(OUT_DIR, exist_ok=True)
out_wb = Workbook()
out_ws = out_wb.active
out_ws.title = "Q2 Solution"
out_ws.append(["Year", "M Age", "F Age", "P(Joint Alive Start)", "P(Failed Year t)", "CF_PV"])
for r in results_term: out_ws.append([r["t"], r["age_m"], r["age_f"], r["p_joint_start"], r["prob_fails"], r["cf_pv"]])
out_ws.append([])
out_ws.append(["EV Term Assurance:", ev_term])
out_ws.append([])
out_ws.append(["t", "P_M(t)", "P_F(t)", "Expected Payment", "PV Payment (Vesting)"])
for t in range(max_years):
    exp_pay = annuity_base if t < 5 else 6000 * (p_m_t[t] + p_f_t[t])
    pv_pay = exp_pay * (v_pay ** t)
    out_ws.append([t, p_m_t[t], p_f_t[t], exp_pay, pv_pay])
out_ws.append([])
out_ws.append(["Value at Vesting:", annuity_value_at_65_M])
out_ws.append(["Prob survive vesting:", prob_both_survive_to_vesting])
out_ws.append(["EV Deferred Annuity:", ev_annuity_total])
workings_path = os.path.join(OUT_DIR, "workings_q2.xlsx")
out_wb.save(workings_path)
