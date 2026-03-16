import math
from openpyxl import load_workbook, Workbook

SOURCE = "exams/CM1B/September_2025/Answer-Booklet.xlsx"
wb = load_workbook(SOURCE, data_only=True)
ws = wb["Q1 Base"]

# Extract values into Python data structures
prices = {}
for row in ws.iter_rows(min_row=5, max_row=24, values_only=True):
    if row[0] is not None:
        n = int(row[0])
        p = row[1]
        if isinstance(p, (int, float)):
            prices[n] = float(p)

f_5_4 = 0.1142 # from row 1
F_8_5 = 0.0872 # from row 2

# Question 1 (i)
# F5,4 = 11.42%. P9 = P5 / (1+f5,4)^4
prices[9] = prices[5] / ((1 + f_5_4) ** 4)

# F8,5 = 8.72%. P13 = P8 * exp(-5 * F8,5)
prices[13] = prices[8] * math.exp(-5 * F_8_5)

print(f"(i) Missing Zero-coupon bond prices:")
print(f"P_9 = {prices[9]:.4f}")
print(f"P_13 = {prices[13]:.4f}")
print()

# Question 1 (ii)
print(f"(ii) Spot rates and 1-year forward rates:")
results = []
for n in range(1, 21):
    y_n = (100.0 / prices[n]) ** (1.0 / n) - 1.0
    if n == 1:
        f_n = (100.0 / prices[1]) - 1.0
    else:
        f_n = (prices[n-1] / prices[n]) - 1.0
    
    results.append({
        "n": n,
        "price": prices[n],
        "spot": y_n,
        "forward": f_n
    })

print(f"{'n':>2} | {'Price':>10} | {'Spot Rate':>10} | {'Forward Rate':>12}")
for r in results:
    print(f"{r['n']:2d} | {r['price']:10.4f} | {r['spot']:10.4%} | {r['forward']:12.4%}")
print()

# Question 1 (iv)
# 5-year par yield
sum_p1_p5 = sum(prices[i]/100.0 for i in range(1, 6))
p5 = prices[5] / 100.0
par_yield_5 = (1.0 - p5) / sum_p1_p5

print(f"(iv) 5-year par yield: {par_yield_5:.4%}")

# Write to workings.xlsx
import os
OUT_DIR = "results/gemini-3.1-pro-high/CM1B/September_2025"
os.makedirs(OUT_DIR, exist_ok=True)
out_wb = Workbook()
out_ws = out_wb.active
out_ws.title = "Q1 Solution"
out_ws.append(["n", "Price", "Spot Rate", "Forward Rate"])
for r in results:
    out_ws.append([r["n"], r["price"], r["spot"], r["forward"]])

out_ws.append([])
out_ws.append(["5-year par yield", par_yield_5])

out_wb.save(os.path.join(OUT_DIR, "workings.xlsx"))
