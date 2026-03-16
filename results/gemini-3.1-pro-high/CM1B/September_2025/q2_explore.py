import openpyxl

wb = openpyxl.load_workbook("exams/CM1B/September_2025/Answer-Booklet.xlsx", data_only=True)
ws = wb["Q2 Base"]

m_term_q = {}
f_term_q = {}
m_ann_q = {}
f_ann_q = {}

for r in range(6, 120):
    # Term assurance
    v_age = ws.cell(row=r, column=1).value
    if v_age is not None and isinstance(v_age, (int, float)):
        age = int(v_age)
        m_term_q[age] = float(ws.cell(row=r, column=2).value)
        f_term_q[age] = float(ws.cell(row=r, column=3).value)
        
    # Annuity
    v_age_ann = ws.cell(row=r, column=5).value
    if v_age_ann is not None and isinstance(v_age_ann, (int, float)):
        age = int(v_age_ann)
        m_ann_q[age] = float(ws.cell(row=r, column=6).value)
        f_ann_q[age] = float(ws.cell(row=r, column=7).value)

print("Term Male ages:", min(m_term_q.keys()), max(m_term_q.keys()))
print("Term Female ages:", min(f_term_q.keys()), max(f_term_q.keys()))
print("Ann Male ages:", min(m_ann_q.keys()), max(m_ann_q.keys()))
print("Ann Female ages:", min(f_ann_q.keys()), max(f_ann_q.keys()))
