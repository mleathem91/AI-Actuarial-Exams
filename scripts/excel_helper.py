"""
Excel Helper Script for LLM Actuarial Exams
============================================

Utility script for reading and describing Excel workbooks used in
computer-based actuarial exam papers (CM1B, CM2B, CS1B, CS2B, etc.).

Usage:
    python scripts/excel_helper.py describe <workbook_path>
    python scripts/excel_helper.py read <workbook_path> <sheet_name>
    python scripts/excel_helper.py read-all <workbook_path>

Commands:
    describe    List all sheets with their dimensions
    read        Print contents of a specific sheet as a formatted table
    read-all    Print contents of all sheets

Requirements:
    pip install openpyxl
"""

import sys
from openpyxl import load_workbook


def describe_workbook(path):
    """List all sheets in a workbook with their dimensions."""
    wb = load_workbook(path, data_only=True)
    print(f"Workbook: {path}")
    print(f"Sheets: {len(wb.sheetnames)}")
    print()
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  {name:30s}  rows={ws.max_row:4d}  cols={ws.max_column:3d}")
    print()


def read_sheet(path, sheet_name, max_col_width=20):
    """Read and print a sheet as a formatted table."""
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found.")
        print(f"Available sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)

    ws = wb[sheet_name]
    print(f"=== {sheet_name} === (rows={ws.max_row}, cols={ws.max_column})")
    print()

    # Read all data
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([format_cell(c) for c in row])

    if not rows:
        print("(empty sheet)")
        return

    # Calculate column widths
    n_cols = max(len(r) for r in rows)
    col_widths = [0] * n_cols
    for row in rows:
        for i, cell in enumerate(row):
            if i < n_cols:
                col_widths[i] = min(max(col_widths[i], len(str(cell))), max_col_width)

    # Print rows
    for i, row in enumerate(rows):
        cells = []
        for j in range(n_cols):
            val = row[j] if j < len(row) else ""
            cells.append(str(val).ljust(col_widths[j])[:max_col_width])
        print(f"R{i+1:3d}: {' | '.join(cells)}")

    print()


def format_cell(value):
    """Format a cell value for display."""
    if value is None:
        return "."
    if isinstance(value, float):
        # Show up to 6 decimal places, strip trailing zeros
        if value == int(value) and abs(value) < 1e10:
            return str(int(value))
        return f"{value:.6f}".rstrip("0").rstrip(".")
    return str(value)


def read_all_sheets(path):
    """Read and print all sheets in a workbook."""
    wb = load_workbook(path, data_only=True)
    for name in wb.sheetnames:
        read_sheet(path, name)
        print()


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    command = sys.argv[1].lower()
    path = sys.argv[2]

    if command == "describe":
        describe_workbook(path)
    elif command == "read":
        if len(sys.argv) < 4:
            print("Error: sheet name required for 'read' command.")
            print("Usage: python excel_helper.py read <workbook_path> <sheet_name>")
            sys.exit(1)
        sheet_name = sys.argv[3]
        read_sheet(path, sheet_name)
    elif command == "read-all":
        read_all_sheets(path)
    else:
        print(f"Unknown command: {command}")
        print("Use: describe, read, or read-all")
        sys.exit(1)


if __name__ == "__main__":
    main()
